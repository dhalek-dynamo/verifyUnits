import openpyxl as xl
import os
import copy

# Forests where all districts are inactive as well - double check these later
INACTIVE_FORESTS = [ 110108, 110105, 110112, 110405, 110608, 110611, 110620, ]
RESULTS_XL = 'results.xlsx'
INPUT_XL = r'Units Comparison as of 2023-11-09 3_58 PM ET (1).xlsx'
PALS = "PALS Prod"
DB_SHEETS = [ 'CARA Test', 'CARA Prod2', 'DataMart Test', 'DataMart Prod', ]
NOPE = 'NOPE'
WORKING = 'WORKING'
VERIFIED = 'Verified'
CHECK = 'Check Further'
UPDATE = 'Update These'
CLEAR_SHEETS = [ VERIFIED, CHECK, UPDATE ]

# Generate a dict of ret[unit_num] = (name, active)
# redundant but oh well
def initial_setup(ws, results):
    pals = dict()
    for rownum in range(6, ws.max_row + 1):
        for c in [2, 6, 8]: #unit num columns for PALS Prod
            unit = ws.cell(rownum, c).value
            name = ws.cell(rownum, c+1).value
            active = ws.cell(rownum, 10).value is None
            if unit in INACTIVE_FORESTS:
                active = False
            pals[unit] = { 'name' : name, 'active' : active }
    return (ws.max_row + 1 - 6), pals

def load_results_ss(results, pals):
    temp = dict(sorted(pals.items())) # just to make sure they're in order
    for unit, val in temp.items():
        results.append([ "", unit, val['name'], val['active']])
    return results
def compare(ws, pals, results):
    for row in range(2, ws.max_row + 1):
        unit = ws.cell(row, 1).value
        results.cell(row, 2).value = unit
        results.cell(row, 3).value = ws.cell(row, 2).value
        if unit in pals:
            if ws.cell(row, 2).value == pals[unit]['name']:
                if ws.cell(row, 3).value == pals[unit]['active']:
                    results.cell(row, 4).value = 'Verified'
                else:
                    results.cell(row, 4).value = 'Check Further'
            else:
                results.cell(row, 4).value = 'Update These'
        else:
            results.cell(row, 4).value = 'Not in PALS'

if __name__ == '__main__':
    wb = xl.load_workbook(INPUT_XL)
    results_xl = xl.load_workbook(RESULTS_XL)
    for s in results_xl.sheetnames:
        results_xl[s].delete_rows(2,results_xl[s].max_row + 1)
    num_rows, pals = initial_setup(wb[PALS], results_xl)

    s = DB_SHEETS[0]
    ws = wb[s]
    print(f'=================\nSheet is: {s} / {ws.title}\n=================')

    current = copy.deepcopy(pals)
    total_checked = 0
    for row in range(2, ws.max_row + 1):
        total_checked += 1
    ws = load_results_ss(results_xl[s], current)
    compare(ws, current, results_xl[s])
    print("Finished")
    count = 0

    results_xl.save(RESULTS_XL)
    print(f'PALS count:        {count + total_checked}')
    print(f'PALS rows:         {num_rows}')
    # print(f'Not in PALS count: {len(not_in_pals)}')
    print( '=======================')
    # print(f'Verified count:    {verified}')


    # for unit, val in pals.items():
    #     pass
    #     count += 1
    #     #if not val[2] else 0


    # for db in bunch-of-dbs ....
    #     do the comparisons here
    # if val[2] is False: # val[2] is flag for it being used, true is good, false means it wasn't in the db
    # print(unit, *val)

    #         t = ws.cell(row, 1).value
    #         t_out = [ ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 4).value]
    #         if t in allPals['four']:
    #             allPals['four'][t][2] = True
    #             if ws.cell(row, 2).value == allPals['four'][t][0] and ws.cell(row, 3).value == allPals['four'][t][1]:
    #                 # verified.append(f'{s} - {row} - {t} - {pals[t][0]}\n')
    #                 wb['VERIFIED'].append(t_out)
    #                 verified += 1
    #             else:
    #                 # add this to NOPE worksheet for now
    #                 wb['NOPE'].append(t_out)
    #                 # print(f"==> NOPE: {ws.cell(row, 1).value} ({ws.cell(row, 2).value}) {ws.cell(row, 3).value}")
    #                 nope += 1
    #         else:
    #             not_in_pals.append(f"==> NOT IN PALS ({s}): {ws.cell(row, 1).value} ({ws.cell(row, 2).value}) {ws.cell(row, 3).value} \n")
    #             wb[WORKING].append(t_out)
    #             # print("=====[ Not in PALS ]=====")

