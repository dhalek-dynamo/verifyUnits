import openpyxl as xl
import os

# Forests where all districts are inactive as well - double check these later
INACTIVE_FORESTS = [ 110108, 110105, 110112, 110405, 110608, 110611, 110620, ]
RESULTS_FILE = 'results.xlsx'
XL_FILE2 = r'Units Comparison 11-09.xlsx'
XL_FILE = r'Units Comparison as of 2023-11-09 3_58 PM ET (1).xlsx'
PALS = "PALS Prod"
CHECK_SHEETS = [ 'CARA Test', 'CARA Prod2', 'DataMart Test', 'DataMart Prod', ]
NOPE = 'NOPE'
WORKING = 'WORKING'
VERIFIED = 'Verified'
CHECK = 'Check Further'
UPDATE = 'Update These'
CLEAR_SHEETS = [ VERIFIED, CHECK, UPDATE ]

# Generate a dict of ret[unit_num] = (name, active)
# redundant but oh well
def initial_setup(ws, results):
    cols = [2, 6, 8] #2?
    ret = dict()
    for rownum in range(6, ws.max_row + 1):
        for c in cols:
            unit = ws.cell(rownum, c).value
            name = ws.cell(rownum, c+1).value
            active = 1 if ws.cell(rownum, 10).value is None else 0
            if (unit):
                    if unit > 1000:
                        ret[unit] = [name, active, False]
                    else:
                        results[CHECK].append([unit, name, active, 'Invalid Unit < 1000'])
    return (ws.max_row + 1 - 6), ret

if __name__ == '__main__':
    wb = xl.load_workbook(XL_FILE)
    results = xl.load_workbook(RESULTS_FILE)

    for s in [WORKING, NOPE, 'VERIFIED']:
        wb[s].delete_rows(1, 5000)
    for s in CLEAR_SHEETS:
        results[s].delete_rows(2,5000)

    num_rows, pals = initial_setup(wb[PALS], results)


    verified = 0
    different = [ ]
    not_in_pals = [ ]
    total_checked = nope = 0

    for s in CHECK_SHEETS:
        print(f'=================\nSheet is: {s} / {s.title}\n=================')
        ws = wb[s]
        for row in range(2, ws.max_row + 1):
            total_checked += 1
            t = ws.cell(row, 1).value
            t_out = [ ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 4).value]
            if t in pals:
                pals[t][2] = True
                if ws.cell(row, 2).value == pals[t][0] and ws.cell(row, 3).value == pals[t][1]:
                    # verified.append(f'{s} - {row} - {t} - {pals[t][0]}\n')
                    wb['VERIFIED'].append(t_out)
                    verified += 1
                else:
                    # add this to NOPE worksheet for now
                    wb['NOPE'].append(t_out)
                    # print(f"==> NOPE: {ws.cell(row, 1).value} ({ws.cell(row, 2).value}) {ws.cell(row, 3).value}")
                    nope += 1
            else:
                not_in_pals.append(f"==> NOT IN PALS ({s}): {ws.cell(row, 1).value} ({ws.cell(row, 2).value}) {ws.cell(row, 3).value} \n")
                wb[WORKING].append(t_out)
                # print("=====[ Not in PALS ]=====")

    print("Finished")
    lv = verified
    lnip = len(not_in_pals)
    count = 0
    # print(*verified)
    # print(*not_in_pals)
    for unit, val in pals.items():
        if val[2] is False:
            print(unit, *val)
        count += 1 if not val[2] else 0

    wb.save(XL_FILE)
    results.save(RESULTS_FILE)
    print(f'Verified count:    {lv}')
    print(f'Not in PALS count: {lnip}')
    print(f'Nope count:        {nope}')
    print(f'PALS count:        {count}')
    print(f'Num_rows:          {num_rows}')
    print(f'Total checked:     {total_checked} / { lv + lnip + nope}')



# for each worksheet remaining:
#   check line by line if row[i], col 0 is in there
#     if so, check name and active
#       if so, add it to a list of 'verified' with sheet name, 3 values
#       else, add it to a list of 'different' with all values from pals, and 3 vals from current table
#     else add it to a list of missing units
#     once i resolve these, i'll generate scripts to handle it
        # print(f'Sheet is: {s} / {ws.title}')
