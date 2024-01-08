import openpyxl as xl
import os
import copy

#todo create a column for sql statements to use for each db results
#todo make sure it's as consistent as possible with db import, maybe put instructions
#     in this file, or a readme, or something

# Assume forests are inactive where all districts are inactive
INACTIVE_FORESTS = [ 110108, 110105, 110112, 110405, 110608, 110611, 110620, ]
RESULTS_XL = 'results.xlsx'
INPUT_XL = r'Units Comparison as of 2023-11-09 3_58 PM ET.xlsx'
PALS = "PALS Prod"
DB_SHEETS = [ 'CARA Test', 'CARA Prod2', 'DataMart Test', 'DataMart Prod', ]
VERIFIED = 'Verified'

def check_worksheet(pals, wb, results):
    for sheet in DB_SHEETS:
        ws = wb[sheet]
        db = dict()
        db_sheet = results[sheet]
        sqlName = 'Name'
        SqlActive = 'Active'
        if 'CARA' in db_sheet.title:
            sqlTable = '[CARA].[dbo].[Unit]'
            sqlId = 'UnitId'
        elif 'DataMart' in db_sheet.title:
            sqlTable = 'datamart1.RefUnits'
            sqlId = 'Id'
        print(db_sheet.title)
        db_sheet.delete_rows(1, 5000)
        db_sheet.append(["Unit Number", "Problem", "Current Name", "CORRECT Name", "isActive", "CORRECT isActive"])
        db_remove_me = [ ]
        for rownum in range(2, ws.max_row+1):
            u = f'{ws.cell(rownum, 1).value}'
            n = ws.cell(rownum, 2).value
            a = True if ws.cell(rownum, 3).value == 1 else False
            db[u] = [n,a]
        unit_missing = pals.keys() - db.keys()
        for unit in db:
            sql = ''
            if unit in pals.keys():
                pals_name = pals[unit][0]
                pals_active = pals[unit][1]
                db_name = db[unit][0]
                db_active = db[unit][1] == 1
                if pals_name == db_name and pals_active == db_active:
                    results[VERIFIED].append([unit, pals_name, pals_active, db_sheet.title])
                else:
                    info = [''] * 7
                    info[0] = unit
                    comma = ''
                    sql = 'UPDATE ' + sqlTable + ' SET '
                    if pals_name != db_name:
                        info[1] = 'Name '
                        info[2] = db_name
                        info[3] = pals_name
                        sql += sqlName + " = '" + pals_name + "'"
                        comma = ', '
                    if pals_active != db_active:
                        info[1] += 'Active'
                        info[4] = db_active
                        info[5] = pals_active
                        sql += comma + SqlActive + " = " + str(pals_active)
                    info[6] = sql + " WHERE " + sqlId + " = " + unit + ";"
                    db_sheet.append(info)
            elif db[unit][1]: #todo inactive ones should be removed, or no? hiding them for now
                db_remove_me.append([unit, "Not in PALS", db[unit][0], db[unit][1]])
        for line in db_remove_me:
            results[ws.title].append(line)
        for unit in unit_missing:
            results[ws.title].append([unit, "In PALS, not here", pals[unit][0], pals[unit][1]])
        results.save(RESULTS_XL)

# Generate a dict of ret[unit_num] = (name, active)
def pals_setup(ws):
    pals = dict()
    for rownum in range(6, ws.max_row + 1):
        for c in [2, 6, 8]: #unit num columns for PALS Prod
            unit = f'{ws.cell(rownum, c).value}'
            name = ws.cell(rownum, c+1).value
            active = ws.cell(rownum, 10).value is None
            if unit and unit in INACTIVE_FORESTS:
                active = False
            pals[unit] = [name, active]
    return pals

def main():
    wb = xl.load_workbook(INPUT_XL)
    results_xl = xl.load_workbook(RESULTS_XL) #todo change to create if not exists
    results_xl[VERIFIED].delete_rows(2,5000)
    pals = pals_setup(wb[PALS]) #load pals dict with the pals (master) data
    check_worksheet(pals, wb, results_xl) # compare the individual db data (worksheets) to the pals data
    print("=====[ DONE ]=====")
    results_xl.save(RESULTS_XL)

if __name__ == '__main__':
    main()
